# coding=utf-8

from math import *

from django.shortcuts import render_to_response
from django.views.decorators.csrf import csrf_exempt
import xlrd
from pymongo import MongoClient
from openpyxl import load_workbook
import sys
from collections import Counter

client = MongoClient('localhost',27017)
db = client.Excel
excel = db.excel

"""
获取经纬度的值并判断经纬度是否一致
H4=设计经度，H5=设计纬度，N4=实测经度，N5=实测纬度，判别门限500m
=IF(OR(H4=0,H4=""),"",
IF(INT(6371004*ACOS((SIN(RADIANS(H5))*SIN(RADIANS(N5))+COS(RADIANS(H5))*COS(RADIANS(N5))*COS(RADIANS(N4-H4)))))<500,"是","否"))
"""


def get_lontitude_latitude(Latitude,Lontitude,measured_sheet,measured_sheet_nrows):
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,8).value) == '' or str(measured_sheet.cell(i,7).value) == '':
            continue
        else:
            global measured_lontitude_value,measured_latitude_value
            measured_lontitude_value = measured_sheet.cell(i,8).value  # 经度和
            measured_latitude_value = measured_sheet.cell(i,7).value  # 纬度和
            break

    error_value = int(6371004 * acos((sin(radians(Latitude)) * sin(radians(measured_latitude_value))
                                      + cos(radians(Latitude)) * cos(radians(measured_latitude_value))
                                      * cos(radians(measured_lontitude_value - Lontitude)))))  # 获取误差值判断经纬度是否一致
    if error_value < 500:
        lontitude_latitude_conclusion = "一致"
    else:
        lontitude_latitude_conclusion = "不一致"

    return [measured_lontitude_value,measured_latitude_value,lontitude_latitude_conclusion,error_value]


"""
获取CGI和LAC的值并判断CGI和归属LAC是否一致
"""


def get_CGI_LAC(CGT_LAC,CGT_CellID,measured_sheet,measured_sheet_nrows):
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,9).value).split('.')[0] != str(CGT_LAC):
            success_LAC = False
        else:
            success_LAC = True
            break
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,10).value).split('.')[0] != str(CGT_CellID):
            success_CellID = False
        else:
            success_CellID = True
            break
    if success_LAC and success_CellID:
        CGI_conclusion = "一致"
    else:
        CGI_conclusion = "不一致"
    if success_LAC:
        LAC_conclusion = "一致"
    else:
        LAC_conclusion = "不一致"

    LAC = []  # 所有的LAC字符串
    for i in range(1,measured_sheet.nrows):
        LAC.append(str(measured_sheet.cell(i,9).value).split('.')[0])
    LAC_list = Counter(LAC).most_common(1)  # [()]
    LAC_most = (LAC_list[0])[0]  # 得到出现次数最多的LAC值

    CellID = []  # 所有的CellID字符串
    for i in range(1,measured_sheet.nrows):
        CellID.append(str(measured_sheet.cell(i,10).value).split('.')[0])
    CellID_list = Counter(CellID).most_common(1)
    CellID_most = (CellID_list[0])[0]  # 得到出现次数最多的CellID值

    measured_CGI = "460-00-" + LAC_most + '-' \
                   + CellID_most
    measured_LAC = LAC_most

    return CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC


"""
获取BCCH的值并判断BCCH是否一致
"""


def get_BCCH(BCCH,measured_sheet,measured_sheet_nrows):
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,11).value).split('.')[0] != str(BCCH).split('.')[0]:
            BCCH_conclusion = "不一致"
        else:
            BCCH_conclusion = "一致"
            break

    list = []  # 所有的BCCH字符串
    for i in range(1,measured_sheet.nrows):
        list.append(str(measured_sheet.cell(i,11).value).split('.')[0])
    BCCH_list = Counter(list).most_common(1)
    BCCH_most = (BCCH_list[0])[0]  # 得到出现次数最多的CellID值
    return BCCH_conclusion,BCCH_most


"""
判断BSIC是否一致
BSIC=NCC+BCC
BSIC = int(design_value.cell(1, 22).value) + int(design_value.cell(1, 13).value)
"""


def get_BSIC(BSIC,measured_sheet,measured_sheet_nrows):
    list = []  # 所有的BCCH字符串
    for i in range(1,measured_sheet.nrows):
        list.append(str(measured_sheet.cell(i,12).value).split('.')[0])
    BSIC_list = Counter(list).most_common(1)
    BSIC_most = (BSIC_list[0])[0]  # 得到出现次数最多的CellID值

    if BSIC_most != str(BSIC).split('.')[0]:
        BSIC_conclusion = "不一致"
    else:
        BSIC_conclusion = "一致"

    return BSIC_conclusion,BSIC_most


"""
判断基站ID是否一致
一致
"""


def get_ID(id,ID,measured_sheet,measured_sheet_nrows):
    list = []  # 所有的CellID字符串
    for i in range(1,measured_sheet.nrows):
        list.append(str(measured_sheet.cell(i,10).value).split('.')[0])
    ID_list = Counter(list).most_common(1)
    ID_most = (ID_list[0])[0]  # 得到出现次数最多的CellID值
    ID_bits = ('%x' % int(ID_most)).zfill(4)
    measured_ID = id + str(ID_bits).upper()
    if measured_ID == ID:
        ID_conclusion = "一致"
    else:
        ID_conclusion = "不一致"

    return measured_ID,ID_conclusion


"""
获取测试点平均电平（dBm)
"""


def get_RxLevelSub_value(measured_sheet,measured_sheet_nrows):
    RxLevelSub = 0
    count = 0
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,15).value) == '':
            continue
        else:
            count = count + 1
            RxLevelSub = RxLevelSub + float(measured_sheet.cell(i,15).value)
    if count == 0:
        RxLevelSub_value = 0
    else:

        RxLevelSub_value = "%.2f" % (RxLevelSub / count)

    return RxLevelSub_value


"""
获取无线接通率（%）和无线掉话率（%）
"""


def get_connected_value_and_unconnected_value(measured_sheet,measured_sheet_nrows):
    count_Call_blocked = 0
    count_Call_attempt = 0
    count_Call_dropped = 0
    count_Call_connected = 0
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,5).value) == "Call blocked":
            count_Call_blocked = count_Call_blocked + 1
        elif str(measured_sheet.cell(i,5).value) == "Call attempt":
            count_Call_attempt = count_Call_attempt + 1
        elif str(measured_sheet.cell(i,5).value) == "Call dropped":
            count_Call_dropped = count_Call_dropped + 1
        elif str(measured_sheet.cell(i,5).value) == "Call connected":
            count_Call_connected = count_Call_connected + 1

    connected = 1 - (count_Call_blocked / count_Call_attempt)
    unconnected = 1 - (count_Call_dropped / count_Call_connected)

    connected_value = "%.2f%%" % (connected * 100)
    unconnected_value = "%.2f%%" % (unconnected * 100)
    return connected_value,unconnected_value


"""
单用户下行峰值吞吐率（kpbs)
"""


def get_max(measured_sheet,measured_sheet_nrows):
    max = 0.00
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,17).value) == '':
            value = 0.00
        else:
            value = float(measured_sheet.cell(i,17).value)
        if value >= max:
            max = value

    return max


"""
语音RxQuality质量0-4级（%）
"""


def var_value(measured_sheet,measured_sheet_nrows):
    count_RxQual = 0
    count_lessfour = 0
    for i in range(1,measured_sheet_nrows):
        if str(measured_sheet.cell(i,16).value) == '':
            continue
        else:
            count_RxQual = count_RxQual + 1
            if float(measured_sheet.cell(i,16).value) < 4:
                count_lessfour = count_lessfour + 1
            else:
                continue
    if count_RxQual == 0:
        value = "100.00%"
    else:
        var = count_lessfour / count_RxQual
        value = "%.2f%%" % (var * 100)
    return value


# 读取excel存放在input_excel_file文件夹下

def input_excel_file(path,file_name):
    try:
        f = open(path,"wb")
        for line in file_name.chunks():
            f.write(line)
        f.close()
    except:
        print "fail"


def submit(request):
    return render_to_response("WebUploader.html")


@csrf_exempt
def testReport(request):
    reload(sys)
    sys.setdefaultencoding('utf-8')

    if request.FILES.get('template_file'):

        template_file = request.FILES.get('template_file')  # 从客户端获取模板excel
        template_path = "./static/" + template_file.name  # 模板excel存储路径
        template_name = template_file.name
        input_excel_file(template_path,template_file)  # 读取模板excel存放在input_excel_file文件夹下
    else:
        return render_to_response("WebUploader.html",{'template_error': True})

    # s1小区的读写
    if request.FILES.get('s1_file'):
        test_1_file = request.FILES.get('s1_file')  # 从客户端获取测试S1的excel
        test_1_path = "./static/" + test_1_file.name  # 测试excel存储路径
        input_excel_file(test_1_path,test_1_file)  # 读取测试excel存放在input_excel_file文件夹下

        s1_name = request.POST['s1_name']  # 从客户端获取小区名用于在模板excel中查询相应小区信息

        measured_data = xlrd.open_workbook(test_1_path,"rb")
        measured_sheet = measured_data.sheet_by_index(0)

        measured_sheet_nrows = measured_sheet.nrows
        s1 = "460-00-" + s1_name
        if excel.find({"CGI": s1}).count() == 0:  # 查询小区是否存在
            return render_to_response("WebUploader.html",{'s1_name_error': True})

        else:
            Lontitude,Latitude,CGI,BCCH,BSIC,ID,LAC,MSC = getMessageFromMongodb(s1_name)  # 获取所要的设计数据

            # 经纬度的实测值和是否一致
            measured_lontitude_value,measured_latitude_value,lontitude_latitude_conclusion,error_value = \
                get_lontitude_latitude(Latitude,Lontitude,measured_sheet,measured_sheet_nrows)
            print measured_lontitude_value,measured_latitude_value,lontitude_latitude_conclusion,error_value

            # CGI和LAC实测值和是否一致
            print CGI
            CGT_LAC = CGI.split('-')[2]
            CGT_CellID = CGI.split('-')[3]
            CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC = get_CGI_LAC(CGT_LAC,CGT_CellID,
                                                                                  measured_sheet,
                                                                                  measured_sheet_nrows)
            print CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC

            # BCCH实测值和是否一致
            BCCH_conclusion,BCCH_most = get_BCCH(BCCH,measured_sheet,measured_sheet_nrows)
            print BCCH_conclusion,BCCH_most

            # BSIC的实测值和是否一致
            BSIC_conclusion,BSIC_most = get_BSIC(BSIC,measured_sheet,measured_sheet_nrows)
            print BSIC_conclusion,BSIC_most

            # 获取测试点平均电平（dBm)
            RxLevelSub_value = get_RxLevelSub_value(measured_sheet, measured_sheet_nrows)
            print RxLevelSub_value

            # 获取无线接通率（%）和无线掉话率（%）
            connected_value,unconnected_value = get_connected_value_and_unconnected_value(measured_sheet,
                                                                                          measured_sheet_nrows)
            print connected_value,unconnected_value

            # 单用户下行峰值吞吐率（kpbs)
            max_value = get_max(measured_sheet,measured_sheet_nrows)
            print max_value

            # 语音RxQuality质量0-4级（%）
            value = var_value(measured_sheet,measured_sheet_nrows)
            print value

            # 判断基站ID是否一致
            id = MSC.split('-')[1]
            measured_ID,ID_conclusion = get_ID(id,ID,measured_sheet,measured_sheet_nrows)
            print measured_ID,ID_conclusion

            writeS1_place(template_path,Lontitude,Latitude,measured_lontitude_value,measured_latitude_value,
                          lontitude_latitude_conclusion,CGI,BCCH,BSIC,ID,LAC,measured_CGI,BCCH_most,BSIC_most,
                          measured_ID,measured_LAC,CGI_conclusion,BCCH_conclusion,BSIC_conclusion,ID_conclusion,
                          LAC_conclusion,RxLevelSub_value,connected_value,unconnected_value,max_value,value
                          )

    else:
        return render_to_response("WebUploader.html",{'test1_error': True})

    # s2 小区的读写
    if request.FILES.get('s2_file'):

        test_2_file = request.FILES.get('s2_file')  # 从客户端获取测试S1的excel
        test_2_path = "./static/" + test_2_file.name  # 测试excel存储路径
        input_excel_file(test_2_path,test_2_file)  # 读取测试excel存放在input_excel_file文件夹下

        s2_name = request.POST['s2_name']  # 从客户端获取小区名用于在模板excel中查询相应小区信息

        measured_data = xlrd.open_workbook(test_2_path,"rb")
        measured_sheet = measured_data.sheet_by_index(0)

        measured_sheet_nrows = measured_sheet.nrows
        s2 = "460-00-" + s2_name
        if excel.find({"CGI": s2}).count() == 0:  # 查询小区是否存在
            return render_to_response("WebUploader.html",{'s2_name_error': True})

        else:
            Lontitude,Latitude,CGI,BCCH,BSIC,ID,LAC,MSC = getMessageFromMongodb(s2_name)  # 获取所要的设计数据

            # CGI和LAC实测值和是否一致

            CGT_LAC = CGI.split('-')[2]
            CGT_CellID = CGI.split('-')[3]
            CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC = get_CGI_LAC(CGT_LAC,CGT_CellID,
                                                                                  measured_sheet,
                                                                                  measured_sheet_nrows)
            print CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC

            # BCCH实测值和是否一致
            BCCH_conclusion,BCCH_most = get_BCCH(BCCH,measured_sheet,measured_sheet_nrows)
            print BCCH_conclusion,BCCH_most

            # BSIC的实测值和是否一致
            BSIC_conclusion,BSIC_most = get_BSIC(BSIC,measured_sheet,measured_sheet_nrows)
            print BSIC_conclusion,BSIC_most

            # 获取测试点平均电平（dBm)
            RxLevelSub_value = get_RxLevelSub_value(measured_sheet,measured_sheet_nrows)
            print RxLevelSub_value

            # 获取无线接通率（%）和无线掉话率（%）
            connected_value,unconnected_value = get_connected_value_and_unconnected_value(measured_sheet,
                                                                                          measured_sheet_nrows)
            print connected_value,unconnected_value

            # 单用户下行峰值吞吐率（kpbs)
            max_value = get_max(measured_sheet,measured_sheet_nrows)
            print max_value

            # 语音RxQuality质量0-4级（%）
            value = var_value(measured_sheet,measured_sheet_nrows)
            print value

            # 判断基站ID是否一致
            id = MSC.split('-')[1]
            measured_ID,ID_conclusion = get_ID(id,ID,measured_sheet,measured_sheet_nrows)
            print measured_ID,ID_conclusion

            writeS2_place(template_path,CGI,BCCH,BSIC,measured_CGI,BCCH_most,BSIC_most,CGI_conclusion,
                          BCCH_conclusion,
                          BSIC_conclusion,RxLevelSub_value,connected_value,unconnected_value,max_value,value)

    # s3 小区的读写
    if request.FILES.get('s3_file'):

        test_3_file = request.FILES.get('s3_file')  # 从客户端获取测试S1的excel
        test_3_path = "./static/" + test_3_file.name  # 测试excel存储路径
        input_excel_file(test_3_path,test_3_file)  # 读取测试excel存放在input_excel_file文件夹下

        s3_name = request.POST['s3_name']  # 从客户端获取小区名用于在模板excel中查询相应小区信息

        measured_data = xlrd.open_workbook(test_3_path,"rb")
        measured_sheet = measured_data.sheet_by_index(0)

        measured_sheet_nrows = measured_sheet.nrows
        s3 = "460-00-" + s3_name

        if excel.find({"CGI": s3}).count() == 0:  # 查询小区是否存在
            return render_to_response("WebUploader.html",{'s3_name_error': True})

        else:
            Lontitude,Latitude,CGI,BCCH,BSIC,ID,LAC,MSC = getMessageFromMongodb(s3_name)  # 获取所要的设计数据

            # CGI和LAC实测值和是否一致

            CGT_LAC = CGI.split('-')[2]
            CGT_CellID = CGI.split('-')[3]
            CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC = get_CGI_LAC(CGT_LAC,CGT_CellID,
                                                                                  measured_sheet,
                                                                                  measured_sheet_nrows)
            print CGI_conclusion,LAC_conclusion,measured_CGI,measured_LAC

            # BCCH实测值和是否一致
            BCCH_conclusion,BCCH_most = get_BCCH(BCCH,measured_sheet,measured_sheet_nrows)
            print BCCH_conclusion,BCCH_most

            # BSIC的实测值和是否一致
            bsic = str(BSIC)
            BSIC_conclusion,BSIC_most = get_BSIC(bsic,measured_sheet,measured_sheet_nrows)
            print BSIC_conclusion,BSIC_most

            # 获取测试点平均电平（dBm)
            RxLevelSub_value = get_RxLevelSub_value(measured_sheet,measured_sheet_nrows)
            print RxLevelSub_value

            # 获取无线接通率（%）和无线掉话率（%）
            connected_value,unconnected_value = get_connected_value_and_unconnected_value(measured_sheet,
                                                                                          measured_sheet_nrows)
            print connected_value,unconnected_value

            # 单用户下行峰值吞吐率（kpbs)
            max_value = get_max(measured_sheet,measured_sheet_nrows)
            print max_value

            # 语音RxQuality质量0-4级（%）
            value = var_value(measured_sheet,measured_sheet_nrows)
            print value

            # 判断基站ID是否一致
            id = MSC.split('-')[1]
            measured_ID,ID_conclusion = get_ID(id,ID,measured_sheet,measured_sheet_nrows)
            print measured_ID,ID_conclusion

            writeS3_place(template_path,CGI,BCCH,BSIC,measured_CGI,BCCH_most,BSIC_most,CGI_conclusion,
                          BCCH_conclusion,BSIC_conclusion,RxLevelSub_value,connected_value,unconnected_value,
                          max_value,value)
    out_template_path = template_file.name
    return render_to_response("WebUploader.html",{"path": out_template_path})


class Write_excel(object):
    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write(self,coord,value):
        # eg: coord:A1
        self.ws.cell(coord).value = value
        self.wb.save(self.filename)

    def return_ws(self):
        return self.wb


# 找到s1小区位置填写

def writeS1_place(template_path,Lontitude,Latitude,measured_lontitude_value,measured_latitude_value,
                  lontitude_latitude_conclusion,CGI,BCCH,BSIC,ID,LAC,measured_CGI,BCCH_most,BSIC_most,
                  measured_ID,measured_LAC,CGI_conclusion,BCCH_conclusion,BSIC_conclusion,ID_conclusion,
                  LAC_conclusion,RxLevelSub_value,connected_value,unconnected_value,max_value,value
                  ):
    wr = Write_excel(filename=template_path)
    wr.write('E16',Lontitude)  # 填写经度设计值
    wr.write('E17',Latitude)  # 填写纬度设计值

    wr.write('H16',measured_lontitude_value)  # 填写经度实测值
    wr.write('H17',measured_latitude_value)  # 填写纬度实测值

    wr.write('K16',lontitude_latitude_conclusion)  # 填写判断经纬度是否一致
    wr.write('K17',lontitude_latitude_conclusion)

    wr.write('E22',CGI)  # 填写CGI设计值
    wr.write('E23',BCCH)  # 填写BCCH设计值
    wr.write('E24',BSIC)  # 填写BSIC设计值
    wr.write('E25',ID)  # 填写基站ID设计值
    wr.write('E26',LAC)  # 填写归属LAC设计值

    wr.write('H22',measured_CGI)  # 填写实测值
    wr.write('H23',BCCH_most)
    wr.write('H24',BSIC_most)
    wr.write('H25',measured_ID)
    wr.write('H26',measured_LAC)

    wr.write('K22',CGI_conclusion)
    wr.write('K23',BCCH_conclusion)
    wr.write('K24',BSIC_conclusion)
    wr.write('K25',ID_conclusion)
    wr.write('K26',LAC_conclusion)

    wr.write('E36',RxLevelSub_value)  # 填写测试点平均电平
    wr.write('E37',connected_value)  # 填写无线接通率
    wr.write('E38',unconnected_value)  # 填写无线掉话率
    wr.write('E39',max_value)  # 填写单用户下行峰值吞吐率（kpbs)
    wr.write('E40',value)  # 填写语音RxQuality质量0-4级（%）


# 找到s2小区填写

def writeS2_place(template_path,CGI,BCCH,BSIC,measured_CGI,BCCH_most,BSIC_most,CGI_conclusion,BCCH_conclusion,
                  BSIC_conclusion,RxLevelSub_value,connected_value,unconnected_value,max_value,value):
    wr = Write_excel(filename=template_path)
    wr.write('F22',CGI)  # 填写CGI设计值
    wr.write('F23',BCCH)  # 填写BCCH设计值
    wr.write('F24',BSIC)  # 填写BSIC设计值

    wr.write('I22',measured_CGI)  # 填写实测值
    wr.write('I23',BCCH_most)
    wr.write('I24',BSIC_most)

    wr.write('K22',CGI_conclusion)  # 有s2小区的话重新填写CGI, BCCH, BSIC是否一致
    wr.write('K23',BCCH_conclusion)
    wr.write('K24',BSIC_conclusion)

    wr.write('K36',RxLevelSub_value)  # 填写测试点平均电平
    wr.write('K37',connected_value)  # 填写无线接通率
    wr.write('K38',unconnected_value)  # 填写无线掉话率
    wr.write('K39',max_value)  # 填写单用户下行峰值吞吐率（kpbs)
    wr.write('K40',value)  # 填写语音RxQuality质量0-4级（%）


# 找到s3小区位置填写
def writeS3_place(template_path,CGI,BCCH,BSIC,measured_CGI,BCCH_most,BSIC_most,CGI_conclusion,BCCH_conclusion,
                  BSIC_conclusion,RxLevelSub_value,connected_value,unconnected_value,max_value,value):
    wr = Write_excel(filename=template_path)
    wr.write('G22',CGI)  # 填写CGI设计值
    wr.write('G23',BCCH)  # 填写BCCH设计值
    wr.write('G24',BSIC)  # 填写BSIC设计值

    wr.write('J22',measured_CGI)  # 填写实测值
    wr.write('J23',BCCH_most)
    wr.write('J24',BSIC_most)

    wr.write('K22',CGI_conclusion)  # 有s2小区的话重新填写CGI, BCCH, BSIC是否一致
    wr.write('K23',BCCH_conclusion)
    wr.write('K24',BSIC_conclusion)

    wr.write('Q36',RxLevelSub_value)  # 填写测试点平均电平
    wr.write('Q37',connected_value)  # 填写无线接通率
    wr.write('Q38',unconnected_value)  # 填写无线掉话率
    wr.write('Q39',max_value)  # 填写单用户下行峰值吞吐率（kpbs)
    wr.write('QK40',value)  # 填写语音RxQuality质量0-4级（%）


def getMessageFromMongodb(village_name):
    cgi = "460-00-" + village_name

    for i in excel.find({"CGI": cgi}):
        global Latitude,Lontitude,CGI,BCCH,BSIC,ID,LAC,MSC
        Lontitude = i[u'经度']
        Latitude = i[u'纬度']
        CGI = i['CGI']
        BCCH = str(i[u'BCCH频点']).split('.')[0]
        BSIC = i['NCC'] * 10 + i['BCC']
        ID = i[u'站号']
        LAC = i['LAC']
        MSC = i[u'所属MSC']
        break

    return Lontitude,Latitude,CGI,BCCH,BSIC,ID,LAC,MSC


def submitMongodbMessage(request):
    return render_to_response("insert.html")


def insertMessageIntoMongodb(request):
    s1 = request.GET['s1_name']
    s2 = request.GET['s2_name']
    s3 = request.GET['s3_name']
    s4 = request.GET['s4_name']
    s5 = request.GET['s5_name']
    s6 = request.GET['s6_name']
    s7 = request.GET['s7_name']
    s8 = request.GET['s8_name']
    s9 = request.GET['s9_name']
    s10 = request.GET['s10_name']
    s11 = request.GET['s11_name']
    s12 = request.GET['s12_name']
    s13 = request.GET['s13_name']
    s14 = request.GET['s14_name']
    s15 = request.GET['s15_name']
    s16 = request.GET['s16_name']
    s17 = request.GET['s17_name']
    s18 = request.GET['s18_name']
    s19 = request.GET['s19_name']
    s20 = request.GET['s20_name']
    s21 = request.GET['s21_name']
    s22 = request.GET['s22_name']
    s23 = request.GET['s23_name']
    s24 = request.GET['s24_name']
    s25 = request.GET['s25_name']
    s26 = request.GET['s26_name']
    s27 = request.GET['s27_name']
    s28 = request.GET['s28_name']
    s29 = request.GET['s29_name']
    s30 = request.GET['s30_name']
    s31 = request.GET['s31_name']
    s32 = request.GET['s32_name']
    s33 = request.GET['s33_name']
    s34 = request.GET['s34_name']
    s35 = request.GET['s35_name']
    s36 = request.GET['s36_name']
    s37 = request.GET['s37_name']
    s38 = request.GET['s38_name']
    s39 = request.GET['s39_name']
    s40 = request.GET['s40_name']
    s41 = request.GET['s41_name']
    s42 = request.GET['s42_name']
    s43 = request.GET['s43_name']
    s44 = request.GET['s44_name']
    s45 = request.GET['s45_name']
    s46 = request.GET['s46_name']
    s47 = request.GET['s47_name']
    data = {
        '小区中文名': s1,'地市': s2,'LAC': s3,'CI': s4,'CGI': s5,
        '站号': s6,'经度': s7,'纬度': s8,'覆盖室内': s9,'900载频数': s10,
        '1800载频数': s11,'AMR是否开': s12,'BBC': s13,'BCCH频点': s14,'TCH频点': s15,
        'TCH总数': s16,'SDCCH信道': s17,'GPRS静态信道数': s18,'GPRS下行功控是否开通': s19,'IRC是否开通': s20,
        'MCCCH是否开通': s21,'NCC': s22,'PDCH静态信道配置数': s23,'开通EDGE载频数': s24,"联合寻呼是否开通": s25,
        '配置的GSM邻区个数': s26,'配置的TD邻区个数': s27,'设备厂商': s28,'所属BSC': s29,'所属BTS': s30,
        '所属MSC': s31,'所用频段': s32,'跳频类型': s33,'载频数量': s34,'支持EGPRS': s35,
        'CCCH信道数': s36,'GPRS动态信道数': s37,'发射功率': s38,'区县': s39,'是否在TD覆盖区内': s40,
        '所属网格编号': s41,'天线挂高': s42,'方位角': s43,'天线下倾角': s44,'电调下倾角': s45,
        '机械下倾角': s46,'区域维护部': s47
    }
    count = excel.find({'小区中文名': s1}).count()
    if s1 == "":
        return render_to_response('insert.html',{'kong_error': True})
    if count != 0:
        return render_to_response('insert.html',{'fail_error': True})
    else:
        excel.insert(data)

        return render_to_response('insert.html',{'success_error': True})
